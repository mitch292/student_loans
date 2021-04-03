from typing import Any, Optional

from django.core.management.base import BaseCommand, CommandParser
from openpyxl import load_workbook
from openpyxl.xml.constants import MIN_ROW

from student_loans.institutions.models import Award, Institution

from ._private import *


class Command(BaseCommand):
    help = (
        "Imports the FSA data for a given quarterly excel that they provide,"
        "found here - https://studentaid.gov/data-center/student/title-iv"
    )

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument("year", type=int, help="The year this data is for.")
        parser.add_argument(
            "quarter", type=int, help="The quarter this data is for. 1 - 4"
        )
        parser.add_argument("file", type=str, help="The path to the file to parse.")

    def handle(self, *args: Any, **options: Any) -> Optional[str]:
        year = options["year"]
        quarter = options["quarter"]
        wb = load_workbook(options["file"])
        ws = wb["Quarterly Activity"]
        for row in ws.iter_rows(DATA_STARTING_ROW):
            (institution, created) = Institution.objects.get_or_create(
                source_id=row[SOURCE_ID_COLUMN].value,
                defaults={
                    "name": row[NAME_COLUMN].value,
                    "state": row[STATE_COLUM].value,
                    "zip": row[ZIP_COLUM].value,
                    "type": map_fsa_str_to_type(row[TYPE_COLUM].value),
                },
            )

            Award.objects.update_or_create(
                institution=institution,
                year=year,
                quarter=quarter,
                loan_type=Award.LoanType.DIRECT_SUBSIDIZED,
                defaults={
                    "receipient_count": row[DL_SUB_RECEPIENTS_COLUMN].value,
                    "loans_originated_count": row[DL_SUB_ORIGINATED_COUNT_COLUMN].value,
                    "loans_originated_amount": row[
                        DL_SUB_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                    "disbursements_count": row[DL_SUB_DISBURSEMENTS_COUNT_COLUMN].value,
                    "disbursements_amount": row[DL_SUB_ORIGINATED_AMOUNT_COLUMN].value,
                },
            )

            Award.objects.update_or_create(
                institution=institution,
                year=year,
                quarter=quarter,
                loan_type=Award.LoanType.DIRECT_UNSUBSIDIZED_UNDERGRADUATE,
                defaults={
                    "receipient_count": row[DL_UNSUB_UNDER_RECEPIENTS_COLUMN].value,
                    "loans_originated_count": row[
                        DL_UNSUB_UNDER_ORIGINATED_COUNT_COLUMN
                    ].value,
                    "loans_originated_amount": row[
                        DL_UNSUB_UNDER_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                    "disbursements_count": row[
                        DL_UNSUB_UNDER_DISBURSEMENTS_COUNT_COLUMN
                    ].value,
                    "disbursements_amount": row[
                        DL_UNSUB_UNDER_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                },
            )

            Award.objects.update_or_create(
                institution=institution,
                year=year,
                quarter=quarter,
                loan_type=Award.LoanType.DIRECT_UNSUBSIDIZED_GRADUATE,
                defaults={
                    "receipient_count": row[DL_UNSUB_GRAD_RECEPIENTS_COLUMN].value,
                    "loans_originated_count": row[
                        DL_UNSUB_GRAD_ORIGINATED_COUNT_COLUMN
                    ].value,
                    "loans_originated_amount": row[
                        DL_UNSUB_GRAD_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                    "disbursements_count": row[
                        DL_UNSUB_GRAD_DISBURSEMENTS_COUNT_COLUMN
                    ].value,
                    "disbursements_amount": row[
                        DL_UNSUB_GRAD_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                },
            )

            Award.objects.update_or_create(
                institution=institution,
                year=year,
                quarter=quarter,
                loan_type=Award.LoanType.DIRECT_PARENT_PLUS,
                defaults={
                    "receipient_count": row[DL_PARENT_PLUS_RECEPIENTS_COLUMN].value,
                    "loans_originated_count": row[
                        DL_PARENT_PLUS_ORIGINATED_COUNT_COLUMN
                    ].value,
                    "loans_originated_amount": row[
                        DL_PARENT_PLUS_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                    "disbursements_count": row[
                        DL_PARENT_PLUS_DISBURSEMENTS_COUNT_COLUMN
                    ].value,
                    "disbursements_amount": row[
                        DL_PARENT_PLUS_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                },
            )

            Award.objects.update_or_create(
                institution=institution,
                year=year,
                quarter=quarter,
                loan_type=Award.LoanType.DIRECT_GRAD_PLUS,
                defaults={
                    "receipient_count": row[DL_GRAD_PLUS_RECEPIENTS_COLUMN].value,
                    "loans_originated_count": row[
                        DL_GRAD_PLUS_ORIGINATED_COUNT_COLUMN
                    ].value,
                    "loans_originated_amount": row[
                        DL_GRAD_PLUS_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                    "disbursements_count": row[
                        DL_GRAD_PLUS_DISBURSEMENTS_COUNT_COLUMN
                    ].value,
                    "disbursements_amount": row[
                        DL_GRAD_PLUS_ORIGINATED_AMOUNT_COLUMN
                    ].value,
                },
            )

            self.stdout.write(
                self.style.SUCCESS(
                    'Successfully Added loan data for "%s"' % institution.name
                )
            )
